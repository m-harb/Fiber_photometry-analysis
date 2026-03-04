from pathlib import Path
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension

from plotting import plot_dff, plot_peth_recording
from pm_analysis import get_recording
from settings import upaths
from utilities import compute_mean_sem


def build_dataframe(recording):
    """
    From a given recording, generates a table that contains the extracted data

    Parameters
    ----------
    recording: dict
        As returned by get_recording

    Returns
    -------
    ev_df:  DataFrame
    """
    mouse_id = recording['metadata']['mouse']
    ev_d = []
    for site in recording['dff'].keys():
        snips = recording['z_snips'].get(site, np.array([]))
        avg, sem = compute_mean_sem(snips)
        resp_amp = recording['resp_amp'].get(site, np.nan)
        resp_auc = recording['resp_auc'].get(site, np.nan)
        act_map = recording['video']['s_map'][site]
        occ_map = recording['video']['occ_map'][site]
        s_occ_map = recording['video']['s_occ_map'][site]
        time_spent = {f'time_{roi}': time for roi,
                      time in recording['video']['time_spent'].items()}
        roi_dff = {f'signal_{roi}': signal for roi,
                   signal in recording['video']['roi_dff'][site].items()}

        c_data_dict = {'snips': snips, 'site': site, 'mouse': str(mouse_id),
                       'events': recording['events'],
                       't_snips': recording['t_snips'],
                       'avg_peth': avg, 'sem_peth': sem,
                       'resp_amp': resp_amp, 'resp_auc': resp_auc,
                       'map': act_map, 'occ_map': occ_map, 's_occ_map': s_occ_map}
        c_data_dict.update(time_spent)
        c_data_dict.update(roi_dff)
        ev_d.append(c_data_dict)
    ev_df = pd.DataFrame(ev_d)
    for k, v in recording['metadata'].items():
        ev_df[k] = v
    return ev_df


def simplify_df(full_df: pd.DataFrame, save_path):
    """
    Simplifying the full data frame so that it can be saved to Excel

    Parameters
    ----------
    full_df: DataFrame
        As created by the function batch
    save_path: Path
        Path to the folder in which we save the excel file

    Returns
    -------
    simp_df: DataFrame
        Simplified version of the DataFrame
    """
    columns_drop = ['snips', 'events', 't_snips',
                    'map', 'occ_map', 's_occ_map',]
    simp_df = full_df.drop(columns_drop, axis=1)
    writer = pd.ExcelWriter(save_path / 'extracted_data.xlsx')
    sh_name = 'raw_data'
    simp_df.to_excel(writer, sheet_name=sh_name, index=False, na_rep='NaN')

    ws = writer.sheets[sh_name]
    dim_holder = DimensionHolder(worksheet=ws)

    for col in range(ws.min_column, ws.max_column + 1):
        column_length = len(simp_df.columns[col - 1])
        dim_holder[get_column_letter(col)] = ColumnDimension(
            ws, min=col, max=col, width=column_length * 1.3)

    ws.column_dimensions = dim_holder

    writer.close()
    return simp_df


def add_social_info(full_df, social_path):
    social_info = pd.read_csv(social_path, dtype={'mouse': str})
    social_info['task'] = 'Social'
    m_df = pd.merge(full_df, social_info, how='left', left_on=['mouse', 'genotype', 'injection', 'task'],
                    right_on=['mouse', 'genotype', 'injection', 'task'])
    m_df['signal_animal_ring'] = np.nan
    m_df['time_animal_ring'] = np.nan
    m_df['signal_empty_ring'] = np.nan
    m_df['time_empty_ring'] = np.nan
    sides = {'left', 'right'}
    for ix_row, row in m_df.iterrows():
        side = row['side']
        o_side = (sides - {side}).pop()
        m_df.loc[ix_row, 'signal_animal_ring'] = m_df.loc[ix_row,
                                                          f'signal_{side}_ring']
        m_df.loc[ix_row, 'time_animal_ring'] = m_df.loc[ix_row,
                                                        f'time_{side}_ring']
        m_df.loc[ix_row, 'signal_empty_ring'] = m_df.loc[ix_row,
                                                         f'signal_{o_side}_ring']
        m_df.loc[ix_row, 'time_empty_ring'] = m_df.loc[ix_row,
                                                       f'time_{o_side}_ring']
    m_df['time_social'] = m_df['time_animal_ring'] + m_df['time_empty_ring']
    return m_df


def batch(data_folder_path, site_names, analysis_path, batch_num=(1, 2), quality_th=2, bin_spacing=5,
          quality_path=None, make_plots=False, force=False):
    """
    Run all the analysis steps on all the recordings in a folder.

    Parameters
    ----------
    data_folder_path: str or Path
        Path to the folder containing recordings
    site_names: dict
        Correspondence between the region names from Neurophotometrics (Region0G, Region1G) and brain region names
    analysis_path: str or Path
        Path to save the intermediate results
    batch_num: tuple of int
        To select animals based on their batch number
    quality_th: int
        Quality threshold to select animals on. 0: no signal, 1: weak, 2: medium, 3: good
    bin_spacing: float
        For the activity and occupancy maps, the bin size in mm
    quality_path: Path, optional
        Path to the CSV file that contains quality estimates for each recording
    make_plots: bool
        Shall we generate the figures
    force: bool
        Shall we re-run the analysis even if it is already done, ignoring previously saved files

    Returns
    -------
    full_df: DataFrame
        Also saved as an HDF5 file
    """
    data_folder_path = Path(data_folder_path)
    analysis_path = Path(analysis_path)
    data_base_name = '_'.join(data_folder_path.parts[-3:])
    h5_name = f'{data_base_name}_quality-{quality_th}_spacing-{bin_spacing}.h5'
    h5_path = analysis_path / h5_name
    if h5_path.exists() and not force:
        full_df = pd.read_hdf(h5_path, key='pm_data')
        return full_df
    l_dfs = []
    # We go through all the video files we can find in this folder and subfolders
    pb = tqdm(list(data_folder_path.rglob('*.avi')))
    for v_path in pb:
        pb.set_description(
            f'analyzing {v_path.parent.parent.parent.name} {v_path.parent.parent.name} {v_path.parent.name} {v_path.name}')
        folder_path = v_path.parent
        recording = get_recording(folder_path, site_names, analysis_path, batch_num=batch_num, quality_th=quality_th,
                                  bin_spacing=bin_spacing, quality_path=quality_path, make_plots=make_plots, force=force)
        if len(recording['dff']) == 0:
            continue
        quality_path = recording['metadata']['quality_path']
        c_df = build_dataframe(recording)
        l_dfs.append(c_df)
        if make_plots:
            plt.close('all')
    full_df = pd.concat(l_dfs)
    full_df.reset_index(drop=True, inplace=True)
    full_df.to_hdf(h5_path, key='pm_data')
    _ = simplify_df(full_df, analysis_path)
    return full_df


if __name__ == '__main__' and True:
    sites = {'Region0G': 'BNST', 'Region1G': 'CeA'}
    df = batch(upaths['data'] / 'OFT', sites, upaths['analysis'], bin_spacing=20,
               force=True, make_plots=True)
    # simplify_df(df, upaths['analysis'])
    
